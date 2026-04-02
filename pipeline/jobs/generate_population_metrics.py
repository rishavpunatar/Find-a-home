from __future__ import annotations

import io
import json
import math
import re
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urljoin
from zoneinfo import ZoneInfo

import openpyxl
import requests
from pyproj import Transformer
from shapely.geometry import Point, shape
from shapely.ops import transform
from shapely.strtree import STRtree

from pipeline.jobs.osrm_utils import fetch_text_with_curl_fallback
from pipeline.jobs.station_scope import candidate_scope_stations


ROOT = Path(__file__).resolve().parents[2]
OUTPUT_PATH = ROOT / 'data' / 'raw' / 'population_metrics.json'
SOURCE_METADATA_PATH = ROOT / 'data' / 'raw' / 'source_metadata.json'
CACHE_DIR = ROOT / 'data' / 'cache' / 'population'

ONS_DATASET_PAGE_URL = (
    'https://www.ons.gov.uk/peoplepopulationandcommunity/populationandmigration/'
    'populationestimates/datasets/lowersuperoutputareamidyearpopulationestimatesnationalstatistics'
)
LSOA_BOUNDARY_QUERY_URL = (
    'https://services1.arcgis.com/ESMARspQHYMw9BZ9/arcgis/rest/services/'
    'LSOA_2021_EW_BFE_V10_RUC/FeatureServer/3/query'
)

STATION_WALK_RADIUS_METERS = 800.0
CRIME_CATCHMENT_RADIUS_METERS = 1800.0
BOUNDARY_QUERY_MARGIN_DEGREES = 0.05
BOUNDARY_BATCH_SIZE = 500

WGS84_TO_OSGB = Transformer.from_crs('EPSG:4326', 'EPSG:27700', always_xy=True)

XLSX_LINK_PATTERN = re.compile(r'href="([^"]+sapelsoabroadage\d+\.xlsx[^"]*)"')
RELEASE_DATE_PATTERN = re.compile(r'Release date:\s*</dt>\s*<dd[^>]*>\s*([^<]+)\s*</dd>', re.IGNORECASE)
SHEET_PATTERN = re.compile(r'^Mid-(20\d{2}) LSOA 2021$')


def write_json(path: Path, payload: Any) -> None:
    path.write_text(f'{json.dumps(payload, indent=2)}\n', encoding='utf-8')


def cache_file_path(name: str) -> Path:
    return CACHE_DIR / name


def download_bytes(url: str, cache_name: str) -> bytes:
    cache_path = cache_file_path(cache_name)
    if cache_path.exists():
        return cache_path.read_bytes()

    cache_path.parent.mkdir(parents=True, exist_ok=True)
    response = requests.get(url, timeout=180)
    response.raise_for_status()
    cache_path.write_bytes(response.content)
    return response.content


def latest_population_workbook() -> tuple[bytes, str, str]:
    page_html = fetch_text_with_curl_fallback(
        ONS_DATASET_PAGE_URL,
        timeout_seconds=45,
        cache_namespace='ons-lsoa-population-page',
        cache_ttl_hours=24 * 14,
    )
    if not page_html:
        raise RuntimeError('Unable to fetch ONS LSOA population dataset page')

    links = XLSX_LINK_PATTERN.findall(page_html)
    if not links:
        raise RuntimeError('Unable to find an XLSX download link on the ONS LSOA population dataset page')

    workbook_url = urljoin('https://www.ons.gov.uk', links[0])
    release_match = RELEASE_DATE_PATTERN.search(page_html)
    release_text = release_match.group(1).strip() if release_match else ''
    release_date = (
        datetime.strptime(release_text, '%d %B %Y').date().isoformat()
        if release_text
        else datetime.now(ZoneInfo('Europe/London')).date().isoformat()
    )

    cache_name = Path(workbook_url.split('?', 1)[0]).name or 'lsoa_population.xlsx'
    return download_bytes(workbook_url, cache_name), workbook_url, release_date


def latest_population_sheet(workbook: openpyxl.Workbook) -> tuple[Any, int]:
    candidates: list[tuple[int, str]] = []
    for sheet_name in workbook.sheetnames:
        match = SHEET_PATTERN.match(sheet_name)
        if not match:
            continue
        candidates.append((int(match.group(1)), sheet_name))

    if not candidates:
        raise RuntimeError('No LSOA population sheets found in the ONS workbook')

    year, sheet_name = max(candidates, key=lambda item: item[0])
    return workbook[sheet_name], year


def lsoa_population_totals(workbook_bytes: bytes) -> tuple[dict[str, float], int]:
    workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes), data_only=True, read_only=True)
    sheet, year = latest_population_sheet(workbook)

    rows = sheet.iter_rows(min_row=4, values_only=True)
    header = next(rows)
    header_names = [str(value).strip() if value is not None else '' for value in header]
    code_index = header_names.index('LSOA 2021 Code')
    total_index = header_names.index('Total')

    output: dict[str, float] = {}
    for row in rows:
        if row is None:
            continue
        raw_code = row[code_index]
        raw_total = row[total_index]
        if not isinstance(raw_code, str) or not raw_code.strip():
            continue
        if not isinstance(raw_total, (int, float)):
            continue
        output[raw_code.strip()] = float(raw_total)

    if not output:
        raise RuntimeError('No LSOA population rows were parsed from the ONS workbook')

    return output, year


def station_bbox() -> tuple[float, float, float, float]:
    stations = candidate_scope_stations()
    lats = [station.coordinate.lat for station in stations]
    lons = [station.coordinate.lon for station in stations]
    return (
        min(lons) - BOUNDARY_QUERY_MARGIN_DEGREES,
        min(lats) - BOUNDARY_QUERY_MARGIN_DEGREES,
        max(lons) + BOUNDARY_QUERY_MARGIN_DEGREES,
        max(lats) + BOUNDARY_QUERY_MARGIN_DEGREES,
    )


def candidate_lsoa_ids(bbox: tuple[float, float, float, float]) -> list[int]:
    west, south, east, north = bbox
    payload = requests.get(
        LSOA_BOUNDARY_QUERY_URL,
        params={
            'f': 'json',
            'where': '1=1',
            'returnIdsOnly': 'true',
            'geometry': f'{west},{south},{east},{north}',
            'geometryType': 'esriGeometryEnvelope',
            'inSR': '4326',
            'spatialRel': 'esriSpatialRelIntersects',
        },
        timeout=60,
    ).json()
    object_ids = payload.get('objectIds')
    if not isinstance(object_ids, list) or not object_ids:
        raise RuntimeError('No LSOA boundary object IDs returned for the station study area')
    return [int(object_id) for object_id in object_ids]


def lsoa_boundary_features(object_ids: list[int]) -> list[dict[str, Any]]:
    features: list[dict[str, Any]] = []
    for start in range(0, len(object_ids), BOUNDARY_BATCH_SIZE):
        batch = object_ids[start : start + BOUNDARY_BATCH_SIZE]
        response = requests.post(
            LSOA_BOUNDARY_QUERY_URL,
            data={
                'f': 'geojson',
                'objectIds': ','.join(str(object_id) for object_id in batch),
                'outFields': 'LSOA21CD,LSOA21NM',
                'returnGeometry': 'true',
                'outSR': '4326',
            },
            timeout=120,
        )
        response.raise_for_status()
        payload = response.json()
        batch_features = payload.get('features')
        if not isinstance(batch_features, list):
            continue
        features.extend(batch_features)
    if not features:
        raise RuntimeError('No LSOA boundary features returned for the station study area')
    return features


def build_lsoa_geometries(
    population_by_lsoa: dict[str, float],
    boundary_features: list[dict[str, Any]],
) -> tuple[list[str], list[Any], list[float], STRtree]:
    codes: list[str] = []
    geometries: list[Any] = []
    populations: list[float] = []

    for feature in boundary_features:
        properties = feature.get('properties')
        geometry = feature.get('geometry')
        if not isinstance(properties, dict) or not isinstance(geometry, dict):
            continue
        code = properties.get('LSOA21CD')
        if not isinstance(code, str):
            continue
        population = population_by_lsoa.get(code)
        if population is None:
            continue

        polygon = transform(WGS84_TO_OSGB.transform, shape(geometry))
        if polygon.is_empty or polygon.area <= 0:
            continue

        codes.append(code)
        geometries.append(polygon)
        populations.append(float(population))

    if not geometries:
        raise RuntimeError('No intersectable LSOA geometries were built for the study area')

    return codes, geometries, populations, STRtree(geometries)


def intersected_population(circle: Any, geometries: list[Any], populations: list[float], tree: STRtree) -> float:
    indexes = tree.query(circle)
    total = 0.0
    for index in indexes:
        geometry = geometries[int(index)]
        overlap_area = geometry.intersection(circle).area
        if overlap_area <= 0:
            continue
        total += populations[int(index)] * (overlap_area / geometry.area)
    return total


def generate_population_metrics() -> dict[str, dict[str, Any]]:
    workbook_bytes, workbook_url, release_date = latest_population_workbook()
    population_by_lsoa, population_year = lsoa_population_totals(workbook_bytes)
    bbox = station_bbox()
    object_ids = candidate_lsoa_ids(bbox)
    features = lsoa_boundary_features(object_ids)
    _codes, geometries, populations, tree = build_lsoa_geometries(population_by_lsoa, features)

    output: dict[str, dict[str, Any]] = {}
    stations = candidate_scope_stations()
    for station in stations:
        x, y = WGS84_TO_OSGB.transform(station.coordinate.lon, station.coordinate.lat)
        station_point = Point(x, y)
        walk_population = intersected_population(
            station_point.buffer(STATION_WALK_RADIUS_METERS),
            geometries,
            populations,
            tree,
        )
        crime_population = intersected_population(
            station_point.buffer(CRIME_CATCHMENT_RADIUS_METERS),
            geometries,
            populations,
            tree,
        )
        output[station.station_code] = {
            'population_in_reference_zone': round(walk_population),
            'population_in_crime_catchment': round(crime_population),
            'status': 'available',
            'confidence': 0.9,
            'provenance': 'direct',
            'methodology_note': (
                'Direct population denominator from ONS mid-year LSOA population estimates intersected with '
                'official 2021 LSOA boundaries. The reference-zone denominator is the population inside the '
                f'station-centred {int(STATION_WALK_RADIUS_METERS)}m walk catchment; the crime denominator uses the '
                f'{int(CRIME_CATCHMENT_RADIUS_METERS)}m crime catchment. Populations are apportioned by the share of '
                'each intersecting LSOA polygon area inside the station buffer.'
            ),
            'source_year': population_year,
            'source_release_date': release_date,
            'source_workbook': workbook_url,
        }

    source_metadata: dict[str, Any] = {}
    if SOURCE_METADATA_PATH.exists():
        existing = json.loads(SOURCE_METADATA_PATH.read_text(encoding='utf-8'))
        if isinstance(existing, dict):
            source_metadata = existing
    source_metadata['population'] = {
        'source': 'ONS mid-year LSOA population estimates + ONS 2021 LSOA boundaries',
        'referencePeriod': (
            f'ONS mid-{population_year} LSOA population estimates apportioned into station '
            f'{int(STATION_WALK_RADIUS_METERS)}m and {int(CRIME_CATCHMENT_RADIUS_METERS)}m buffers'
        ),
        'releaseDate': release_date,
    }
    write_json(SOURCE_METADATA_PATH, source_metadata)
    return output


def main() -> None:
    payload = generate_population_metrics()
    write_json(OUTPUT_PATH, payload)
    print(f'Generated population metrics for {len(payload)} stations -> {OUTPUT_PATH}')


if __name__ == '__main__':
    main()
